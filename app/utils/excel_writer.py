from __future__ import annotations
import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.models import BOMRow

_GREEN = PatternFill("solid", fgColor="C6EFCE")
_YELLOW = PatternFill("solid", fgColor="FFEB9C")
_RED = PatternFill("solid", fgColor="FFC7CE")
_GRAY = PatternFill("solid", fgColor="D9D9D9")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")

_ADDED_COLS = [
    "Best Supplier",
    "Best Unit Price (USD)",
    "Best Unit Price (VND)",
    "Total Line (USD)",
    "Total Line (VND)",
    "Mouser Price (USD)",
    "DigiKey Price (USD)",
    "LCSC Price (USD)",
]

_THIN = Side(style="thin", color="AAAAAA")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _fmt_usd(v: float | None) -> str:
    return f"${v:.4f}" if v is not None else "N/A"


def _fmt_vnd(v: float | None) -> str:
    if v is None:
        return "N/A"
    return f"{v:,.0f} ₫"


def write_excel(rows: list[BOMRow], usd_to_vnd: float, job_id: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM Pricing"

    orig_cols = list(rows[0].original.keys()) if rows else []
    all_cols = orig_cols + _ADDED_COLS

    # Header row
    for col_idx, col_name in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = _HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    ws.row_dimensions[1].height = 30

    grand_total_usd = 0.0
    grand_total_vnd = 0.0

    for row_idx, bom_row in enumerate(rows, 2):
        # Determine fill color
        if bom_row.no_part_number:
            fill = _GRAY
        else:
            found_count = sum(
                1 for r in [bom_row.mouser, bom_row.digikey, bom_row.lcsc] if r is not None
            )
            if found_count == 0:
                fill = _RED
            elif found_count == 1:
                fill = _YELLOW
            else:
                fill = _GREEN

        # Write original columns
        for col_idx, col_name in enumerate(orig_cols, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=bom_row.original.get(col_name))
            cell.fill = fill
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center")

        # Write added columns
        base = len(orig_cols) + 1

        if bom_row.no_part_number:
            added_values = ["No Part Number", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A"]
        else:
            best = bom_row.best
            best_usd = best.unit_price_usd if best else None
            best_vnd = (best_usd * usd_to_vnd) if best_usd is not None else None
            qty = bom_row.quantity
            total_usd = (best_usd * qty) if best_usd is not None else None
            total_vnd = (total_usd * usd_to_vnd) if total_usd is not None else None

            if total_usd is not None:
                grand_total_usd += total_usd
            if total_vnd is not None:
                grand_total_vnd += total_vnd

            added_values = [
                best.supplier if best else "Not Found",
                _fmt_usd(best_usd),
                _fmt_vnd(best_vnd),
                _fmt_usd(total_usd),
                _fmt_vnd(total_vnd),
                _fmt_usd(bom_row.mouser.unit_price_usd if bom_row.mouser else None),
                _fmt_usd(bom_row.digikey.unit_price_usd if bom_row.digikey else None),
                _fmt_usd(bom_row.lcsc.unit_price_usd if bom_row.lcsc else None),
            ]
        for i, val in enumerate(added_values):
            cell = ws.cell(row=row_idx, column=base + i, value=val)
            cell.fill = fill
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Grand total row
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value="GRAND TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=1).fill = _TOTAL_FILL
    ws.cell(row=total_row, column=1).border = _BORDER

    total_usd_col = len(orig_cols) + _ADDED_COLS.index("Total Line (USD)") + 1
    total_vnd_col = len(orig_cols) + _ADDED_COLS.index("Total Line (VND)") + 1

    cell_usd = ws.cell(row=total_row, column=total_usd_col, value=_fmt_usd(grand_total_usd))
    cell_usd.font = Font(bold=True)
    cell_usd.fill = _TOTAL_FILL
    cell_usd.border = _BORDER
    cell_usd.alignment = Alignment(horizontal="center")

    cell_vnd = ws.cell(row=total_row, column=total_vnd_col, value=_fmt_vnd(grand_total_vnd))
    cell_vnd.font = Font(bold=True)
    cell_vnd.fill = _TOTAL_FILL
    cell_vnd.border = _BORDER
    cell_vnd.alignment = Alignment(horizontal="center")

    # Exchange rate note
    note_row = total_row + 1
    ws.cell(row=note_row, column=1, value=f"Exchange rate: 1 USD = {usd_to_vnd:,.0f} VND")
    ws.cell(row=note_row, column=1).font = Font(italic=True, color="666666")

    # Auto-fit column widths
    for col_idx in range(1, len(all_cols) + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)

    path = os.path.join(tempfile.gettempdir(), f"{job_id}.xlsx")
    wb.save(path)
    return path
